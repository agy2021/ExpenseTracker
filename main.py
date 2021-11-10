from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import datetime

wb_name = "data.xlsx"
wb = load_workbook(wb_name)
ws = wb.active

date_and_time = datetime.datetime.now().strftime("%H-%M-%S-%d-%m-%Y")

f = open("file.txt", "r")
val = int(f.read())

first = input("(1)Add expense, (2)delete all expenses, or press \"Enter/return\" to view expenses. ")

if first == "":
    for row in range(1, val):
        for col in range(1, 4):
            char = get_column_letter(col)
            print(ws[char + str(row)].value, "\n")

    exit()
elif first == "2":
    warning = input("Are you sure you would like to delete all expenses? (y/n) ")
    if warning == "y":
        print("Deleting all expenses...")
        ws.delete_cols(1, 3)
        wb.save(wb_name)
        print("All expenses deleted.")
        exit()
    else:
        print("You have chosen not to delete all expenses.")
        exit()

name = input("Enter name of what was bought: ")
price = input("Enter price of: " + name + " here: ")
price = "$" + price

name = name + "\n"

date_and_time = date_and_time + "\n"

print("Saving...")
ws.append([name, price, date_and_time])
wb.save(wb_name)
print("Saved!")
print("\nRemember, the date that is saved to the Excel spreadsheet is: Hour-Minute-Second-Day-Month-Year")